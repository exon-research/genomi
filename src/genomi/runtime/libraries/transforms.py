"""Post-download processing steps for the central library manager.

Each ``Transform`` in ``spec.py`` maps to a function here. The manager downloads
raw bytes via ``source_fetch`` then applies the spec's transform before the
source counts as installed. This module owns the file-level mechanics that used
to live in ``static_dependencies`` (FASTA gunzip + ``.fai``), the installer
script (CellMarker XLSX→TSV, tarball extraction, sha256), so there is one place
that knows how a downloaded artifact becomes an installed library.
"""

from __future__ import annotations

import contextlib
import csv
import gzip
import hashlib
import os
import shutil
import tarfile
import time
from pathlib import Path


class file_lock:
    """Cross-process exclusive lock via O_EXCL create, used to serialize the
    multi-GB reference-FASTA download/transform so concurrent runs don't race."""

    def __init__(self, path: Path, *, timeout_seconds: int = 7200) -> None:
        self.path = path
        self.timeout_seconds = timeout_seconds
        self.fd: int | None = None

    def __enter__(self) -> "file_lock":
        started = time.monotonic()
        while True:
            try:
                self.fd = os.open(self.path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                os.write(self.fd, str(os.getpid()).encode("ascii"))
                return self
            except FileExistsError as exc:
                if time.monotonic() - started > self.timeout_seconds:
                    raise TimeoutError(f"timed out waiting for library lock: {self.path}") from exc
                time.sleep(2)

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
        if self.fd is not None:
            os.close(self.fd)
            self.fd = None
        with contextlib.suppress(FileNotFoundError):
            self.path.unlink()


def sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def verify_sha256(path: Path, expected: str) -> None:
    """Raise ``ValueError`` if ``path`` does not hash to ``expected``."""
    actual = sha256_file(path)
    if actual.lower() != expected.lower():
        raise ValueError(f"checksum mismatch for {path}: expected {expected}, got {actual}")


def gunzip_faidx(gz_path: Path, fasta_out: Path, fai_out: Path) -> None:
    """Decompress a ``.fa.gz`` to ``fasta_out`` and build its ``.fai`` index.

    Used for the reference FASTA libraries. Writes through a temp file so a
    crash never leaves a half-written FASTA in place.
    """
    fasta_out.parent.mkdir(parents=True, exist_ok=True)
    tmp = fasta_out.with_name(fasta_out.name + ".gunzip.tmp")
    try:
        with gzip.open(gz_path, "rb") as source, tmp.open("wb") as target:
            shutil.copyfileobj(source, target)
        tmp.replace(fasta_out)
    finally:
        if tmp.exists():
            tmp.unlink()
    write_fasta_index(fasta_out, fai_out)


def write_fasta_index(fasta: Path, output: Path) -> None:
    rows: list[str] = []
    current_name: str | None = None
    current_length = 0
    current_offset = 0
    line_bases = 0
    line_width = 0

    def flush() -> None:
        nonlocal current_name, current_length, current_offset, line_bases, line_width
        if current_name is None:
            return
        rows.append(f"{current_name}\t{current_length}\t{current_offset}\t{line_bases}\t{line_width}\n")

    with fasta.open("rb") as handle:
        while True:
            offset = handle.tell()
            line = handle.readline()
            if not line:
                break
            if line.startswith(b">"):
                flush()
                current_name = line[1:].split(None, 1)[0].decode("ascii", errors="replace")
                current_length = 0
                current_offset = handle.tell()
                line_bases = 0
                line_width = 0
                continue
            bases = line.rstrip(b"\r\n")
            if current_name is None or not bases:
                continue
            if line_bases == 0:
                line_bases = len(bases)
                line_width = len(line)
            current_length += len(bases)
            if current_offset == 0:
                current_offset = offset
        flush()
    output.write_text("".join(rows), encoding="utf-8")


def xlsx_to_tsv(source: Path, output: Path) -> None:
    """Normalize the CellMarker 2.0 human XLSX export to a Genomi marker TSV."""
    from openpyxl import load_workbook

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]
    header_index = {name.lower(): index for index, name in enumerate(header)}

    def value(row: tuple[object, ...], name: str) -> str:
        index = header_index.get(name.lower())
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    tmp = output.with_name(output.name + ".partial")
    with tmp.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "cell_type",
                "gene_symbol",
                "marker",
                "lineage_context",
                "marker_strength",
                "record_id",
                "reference",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        for row_number, row in enumerate(rows, start=2):
            cell_type = value(row, "cell_name")
            gene_symbol = value(row, "Symbol") or value(row, "marker")
            if not cell_type or not gene_symbol:
                continue
            writer.writerow(
                {
                    "cell_type": cell_type,
                    "gene_symbol": gene_symbol,
                    "marker": value(row, "marker"),
                    "lineage_context": value(row, "tissue_type") or value(row, "tissue_class"),
                    "marker_strength": value(row, "cancer_type") or value(row, "cell_type"),
                    "record_id": value(row, "cellontology_id") or f"CellMarker2.0:{row_number}",
                    "reference": "https://bio-bigdata.hrbmu.edu.cn/CellMarker/",
                }
            )
    tmp.replace(output)


def extract_named_binary(tarball: Path, install_dir: Path, binary_name: str, *, compression: str = "bz2") -> Path:
    """Extract a single named binary from an aligner release tarball.

    The minimap2 / bwa-mem2 release tarballs contain a top-level directory with
    the binary at its root; this flattens that and writes the binary directly to
    ``install_dir/binary_name``, returning its path.
    """
    install_dir.mkdir(parents=True, exist_ok=True)
    extracted: Path | None = None
    with tarfile.open(tarball, f"r:{compression}") as tar:
        for member in tar.getmembers():
            if not member.isfile() or Path(member.name).name != binary_name:
                continue
            member.name = binary_name  # flatten the leading release dir
            tar.extract(member, install_dir, filter="data")
            extracted = install_dir / binary_name
            break
    if extracted is None or not extracted.is_file():
        raise ValueError(f"tarball {tarball} did not contain a {binary_name} binary")
    extracted.chmod(0o755)
    return extracted


def extract_flat_tarball(tarball: Path, target_dir: Path, *, compression: str = "gz") -> list[str]:
    """Extract every file from a tarball into ``target_dir``, flattening any
    leading directory. Used for the ancestry reference panel. Returns the list
    of extracted file names."""
    target_dir.mkdir(parents=True, exist_ok=True)
    names: list[str] = []
    with tarfile.open(tarball, f"r:{compression}") as tar:
        members = [m for m in tar.getmembers() if m.isfile()]
        if not members:
            raise ValueError(f"tarball is empty: {tarball}")
        for member in members:
            name = Path(member.name).name  # flatten leading panel-NNN/ directory
            if not name:
                continue
            member.name = name
            tar.extract(member, target_dir, filter="data")
            names.append(name)
    return names
