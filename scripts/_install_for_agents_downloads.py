"""Download, verification, and library-install helpers for the Genomi agent
installer.

Split out of ``install_for_agents.py`` to keep each module under the line
budget. Constants and the print helper come from ``_install_for_agents_lib``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tarfile
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from _install_for_agents_lib import (
    ANCESTRY_PANEL_TARBALL_SHA256,
    ANCESTRY_PANEL_TARBALL_URL,
    BWA_MEM2_LINUX_X64_SHA256,
    BWA_MEM2_LINUX_X64_URL,
    BWA_MEM2_VERSION,
    CELLMARKER_HUMAN_URL,
    ENCODE_CCRE_GRCH38_URL,
    GENCODE_GRCH37_URL,
    GENCODE_GRCH38_URL,
    GENOMI_USER_AGENT,
    LIBRARIES,
    LIBRARY_SIZES,
    MINIMAP2_LINUX_X64_SHA256,
    MINIMAP2_LINUX_X64_URL,
    MINIMAP2_VERSION,
    PANGLAODB_MARKERS_URL,
    PHARMCAT_RELEASES_API_URL,
    SRC_DIR,
    print_summary,
)


def _abort_on_existing_install(selected: list[str], *, force: bool) -> None:
    """Refuse to overwrite a populated GENOMI_HOME unless --force was passed.

    Protects an existing install when a second one is run into the same home
    by accident. Only triggers when libraries would actually be installed and
    the home already has populated resource/reference/tools directories.
    """
    if force or not selected:
        return
    home = Path(os.environ.get("GENOMI_HOME") or str(Path("~/.genomi").expanduser())).expanduser()
    if not home.exists():
        return
    populated = []
    for child in ("resources", "reference", "tools"):
        sub = home / child
        if sub.is_dir() and any(sub.iterdir()):
            populated.append(child)
    if not populated:
        return
    print(
        f"Refusing to install into existing GENOMI_HOME {home}: already contains {', '.join(populated)}.\n"
        "Pass --force to refresh selected libraries, or set --genomi-home to a clean path.",
        file=sys.stderr,
    )
    raise SystemExit(2)


def _verify(label: str, command: list[str]) -> None:
    """Run a verify command quietly. Print one OK/FAIL line; on failure dump output.

    Sets GENOMI_CLI=1 so the install-time verify shells past the CLI gate.
    Host agents at runtime should not set this and should use the MCP tools.
    """
    env = {**os.environ, "GENOMI_CLI": "1"}
    result = subprocess.run(command, capture_output=True, text=True, env=env)
    if result.returncode == 0:
        print(f"verify: {label} ok")
        return
    print(f"verify: {label} FAILED (exit {result.returncode})", file=sys.stderr)
    if result.stdout:
        sys.stderr.write(result.stdout)
    if result.stderr:
        sys.stderr.write(result.stderr)
    raise SystemExit(result.returncode)


def install_libraries(libraries: list[str], *, force: bool, args: argparse.Namespace | None = None) -> None:
    if str(SRC_DIR) not in sys.path:
        sys.path.insert(0, str(SRC_DIR))

    from genomi.capabilities.analytical_grounding import (
        analytical_grounding,
    )
    from genomi.capabilities.ancestry.panel_build import (
        build_grch37_panel_from_grch38,
    )
    from genomi.capabilities.ancestry.reference_panels import (
        PANEL_FILES as ANCESTRY_PANEL_FILES,
    )
    from genomi.capabilities.ancestry.reference_panels import (
        panel_dir as ancestry_panel_dir,
    )
    from genomi.capabilities.phenotype.phenotype import (
        GENCC_SUBMISSIONS_URL,
        HPO_DISEASE_ANNOTATION_URL,
        HPO_GENE_ANNOTATION_URL,
        _resolve_gencc_file,
        _resolve_public_annotation_file,
    )
    from genomi.runtime.paths import (
        bwa_mem2_binary_path,
        bwa_mem2_install_dir,
        minimap2_binary_path,
        minimap2_install_dir,
        pharmcat_jar_path,
    )
    from genomi.runtime.liftover import (
        CHAIN_FILES as LIFTOVER_CHAIN_FILES,
        CHAIN_SOURCE_URLS as LIFTOVER_CHAIN_SOURCE_URLS,
        liftover_resources_dir,
    )
    from genomi.runtime.static_dependencies import (
        ensure_reference_fasta,
        ensure_shared_clinvar_vcf,
    )

    for library in libraries:
        disk = LIBRARY_SIZES.get(library, "?")
        print(f"Installing {library}: {LIBRARIES[library]} ({disk})")
        if library == "clinvar-grch38":
            print_summary(ensure_shared_clinvar_vcf(genome_build="GRCh38", force=force))
        elif library == "clinvar-grch37":
            print_summary(ensure_shared_clinvar_vcf(genome_build="GRCh37", force=force))
        elif library == "reference-grch38":
            print_summary(ensure_reference_fasta(genome_build="GRCh38", force=force))
        elif library == "reference-grch37":
            print_summary(ensure_reference_fasta(genome_build="GRCh37", force=force))
        elif library == "hpo":
            gene_file = _resolve_public_annotation_file(
                annotation_file=None,
                annotation_url=HPO_GENE_ANNOTATION_URL,
                cache_name="phenotype_to_genes.txt",
                download_annotations=True,
            )
            disease_file = _resolve_public_annotation_file(
                annotation_file=None,
                annotation_url=HPO_DISEASE_ANNOTATION_URL,
                cache_name="phenotype.hpoa",
                download_annotations=True,
            )
            print_summary(
                {"status": "completed", "output": str(gene_file), "library": "hpo-gene-annotations"}
            )
            print_summary(
                {"status": "completed", "output": str(disease_file), "library": "hpo-disease-annotations"}
            )
        elif library == "gencc":
            path = _resolve_gencc_file(gencc_file=None, gencc_url=GENCC_SUBMISSIONS_URL, download_gencc=True)
            print_summary({"status": "completed", "output": str(path), "library": "gencc"})
        elif library == "gencode-grch38":
            print_summary(
                download_library_file(
                    GENCODE_GRCH38_URL,
                    analytical_grounding.analytical_library_path("gencode-grch38"),
                    library=library,
                    force=force,
                )
            )
        elif library == "gencode-grch37":
            print_summary(
                download_library_file(
                    GENCODE_GRCH37_URL,
                    analytical_grounding.analytical_library_path("gencode-grch37"),
                    library=library,
                    force=force,
                )
            )
        elif library == "encode-ccre-grch38":
            print_summary(
                download_library_file(
                    ENCODE_CCRE_GRCH38_URL,
                    analytical_grounding.analytical_library_path("encode-ccre-grch38"),
                    library=library,
                    force=force,
                )
            )
        elif library == "panglaodb-markers":
            print_summary(
                download_library_file(
                    PANGLAODB_MARKERS_URL,
                    analytical_grounding.analytical_library_path("panglaodb-markers"),
                    library=library,
                    force=force,
                )
            )
        elif library == "cellmarker-human":
            print_summary(
                install_cellmarker_human(
                    analytical_grounding.analytical_library_path("cellmarker-human"),
                    force=force,
                )
            )
        elif library == "msigdb-hallmark":
            print_summary(
                install_msigdb_hallmark(
                    analytical_grounding.analytical_library_path("msigdb-hallmark"),
                    source_path=getattr(args, "msigdb_gmt", None) if args else None,
                    source_url=getattr(args, "msigdb_gmt_url", None) if args else None,
                    force=force,
                )
            )
        elif library == "pharmcat":
            print_summary(
                install_pharmcat(
                    pharmcat_jar_path(),
                    version=getattr(args, "pharmcat_version", None) if args else None,
                    force=force,
                )
            )
        elif library == "minimap2-binary":
            print_summary(
                install_aligner_binary(
                    install_dir=minimap2_install_dir(),
                    binary_path=minimap2_binary_path(),
                    library=library,
                    archive_url=MINIMAP2_LINUX_X64_URL,
                    archive_sha256=MINIMAP2_LINUX_X64_SHA256,
                    binary_name="minimap2",
                    version=MINIMAP2_VERSION,
                    force=force,
                )
            )
        elif library == "bwa-mem2-binary":
            print_summary(
                install_aligner_binary(
                    install_dir=bwa_mem2_install_dir(),
                    binary_path=bwa_mem2_binary_path(),
                    library=library,
                    archive_url=BWA_MEM2_LINUX_X64_URL,
                    archive_sha256=BWA_MEM2_LINUX_X64_SHA256,
                    binary_name="bwa-mem2",
                    version=BWA_MEM2_VERSION,
                    force=force,
                )
            )
        elif library == "ancestry-1000g-30x-grch38":
            target = ancestry_panel_dir()
            prebuilt_dir = getattr(args, "ancestry_panel_dir", None) if args else None
            override_url = getattr(args, "ancestry_panel_url", None) if args else None
            if prebuilt_dir:
                print_summary(_copy_ancestry_panel(prebuilt_dir, target, panel_files=ANCESTRY_PANEL_FILES, force=force))
            else:
                tarball_url = override_url or ANCESTRY_PANEL_TARBALL_URL
                if not tarball_url:
                    raise SystemExit(
                        "No ancestry panel tarball URL is configured. Either pass "
                        "--ancestry-panel-url <https://...> to point at a published release, "
                        "or --ancestry-panel-dir <path> to copy a locally-built panel."
                    )
                print_summary(
                    _download_ancestry_panel(
                        tarball_url=tarball_url,
                        target=target,
                        expected_sha256=ANCESTRY_PANEL_TARBALL_SHA256,
                        force=force,
                    )
                )
        elif library == "liftover-chains":
            target_dir = liftover_resources_dir()
            for filename in sorted(set(LIFTOVER_CHAIN_FILES.values())):
                print_summary(
                    download_library_file(
                        LIFTOVER_CHAIN_SOURCE_URLS[filename],
                        target_dir / filename,
                        library=f"liftover-chains:{filename}",
                        force=force,
                    )
                )
        elif library == "ancestry-1000g-30x-grch37":
            print_summary(build_grch37_panel_from_grch38(force=force))


def _copy_ancestry_panel(
    source_dir: str,
    target: Path,
    *,
    panel_files: tuple[str, ...],
    force: bool,
) -> dict[str, object]:
    source = Path(source_dir).expanduser()
    if not source.is_dir():
        raise SystemExit(f"--ancestry-panel-dir does not exist: {source}")
    missing = [name for name in panel_files if not (source / name).exists()]
    if missing:
        raise SystemExit(
            f"--ancestry-panel-dir is missing panel files: {', '.join(missing)}"
        )
    if target.exists() and force:
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)
    for name in panel_files:
        shutil.copyfile(source / name, target / name)
    return {
        "status": "completed",
        "library": "ancestry-1000g-30x-grch38",
        "manifest_path": str(target / panel_files[0]),
        "source": str(source),
    }


def _download_ancestry_panel(
    *,
    tarball_url: str,
    target: Path,
    expected_sha256: str | None,
    force: bool,
) -> dict[str, object]:
    manifest = target / "manifest.json"
    if manifest.exists() and not force:
        return {
            "status": "cached",
            "library": "ancestry-1000g-30x-grch38",
            "manifest_path": str(manifest),
        }
    target.mkdir(parents=True, exist_ok=True)
    tarball_path = target.with_name(target.name + ".tarball.partial")
    request = urllib.request.Request(
        tarball_url, headers={"User-Agent": GENOMI_USER_AGENT}
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response, tarball_path.open("wb") as handle:
            shutil.copyfileobj(response, handle)
    except urllib.error.HTTPError as exc:
        tarball_path.unlink(missing_ok=True)
        raise SystemExit(
            f"Could not download ancestry panel tarball from {tarball_url}: HTTP {exc.code}. "
            f"If a release does not exist yet, pass --ancestry-panel-dir /path/to/prebuilt or "
            f"--ancestry-panel-url <mirror> to override."
        ) from exc
    if expected_sha256:
        actual = _sha256_file(tarball_path)
        if actual.lower() != expected_sha256.lower():
            tarball_path.unlink(missing_ok=True)
            raise SystemExit(
                f"Ancestry panel tarball checksum mismatch: expected {expected_sha256}, got {actual}."
            )
    with tarfile.open(tarball_path, "r:gz") as tar:
        members = [m for m in tar.getmembers() if m.isfile()]
        if not members:
            raise SystemExit(f"Ancestry panel tarball is empty: {tarball_url}")
        for member in members:
            name = Path(member.name).name  # flatten leading panel-NNN/ directory
            if not name:
                continue
            member.name = name
            tar.extract(member, target, filter="data")
    tarball_path.unlink(missing_ok=True)
    return {
        "status": "completed",
        "library": "ancestry-1000g-30x-grch38",
        "manifest_path": str(manifest),
        "source": tarball_url,
    }


def _sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def download_library_file(url: str, output: Path, *, library: str, force: bool) -> dict[str, object]:
    if output.is_file() and not force:
        return {"status": "cached", "output": str(output), "library": library}
    output.parent.mkdir(parents=True, exist_ok=True)
    tmp = output.with_name(output.name + ".partial")
    request = urllib.request.Request(url, headers={"User-Agent": GENOMI_USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response, tmp.open("wb") as handle:
        shutil.copyfileobj(response, handle)
    tmp.replace(output)
    write_manifest(output, library=library, source_url=url, transform="")
    return {"status": "completed", "output": str(output), "library": library}


def install_cellmarker_human(output: Path, *, force: bool) -> dict[str, object]:
    if output.is_file() and not force:
        return {"status": "cached", "output": str(output), "library": "cellmarker-human"}
    xlsx_path = output.with_suffix(".source.xlsx")
    download_library_file(CELLMARKER_HUMAN_URL, xlsx_path, library="cellmarker-human-source", force=force)
    normalize_cellmarker_xlsx(xlsx_path, output)
    write_manifest(
        output,
        library="cellmarker-human",
        source_url=CELLMARKER_HUMAN_URL,
        transform="CellMarker 2.0 human XLSX normalized to Genomi marker-table TSV.",
    )
    return {"status": "completed", "output": str(output), "library": "cellmarker-human"}


def normalize_cellmarker_xlsx(source: Path, output: Path) -> None:
    from openpyxl import load_workbook

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]
    header_index = {name.lower(): index for index, name in enumerate(header)}

    def value(row: tuple[object, ...], name: str) -> str:
        index = header_index.get(name.lower())
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    tmp = output.with_name(output.name + ".partial")
    with tmp.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "cell_type",
                "gene_symbol",
                "marker",
                "lineage_context",
                "marker_strength",
                "record_id",
                "reference",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        for row_number, row in enumerate(rows, start=2):
            cell_type = value(row, "cell_name")
            gene_symbol = value(row, "Symbol") or value(row, "marker")
            if not cell_type or not gene_symbol:
                continue
            writer.writerow(
                {
                    "cell_type": cell_type,
                    "gene_symbol": gene_symbol,
                    "marker": value(row, "marker"),
                    "lineage_context": value(row, "tissue_type") or value(row, "tissue_class"),
                    "marker_strength": value(row, "cancer_type") or value(row, "cell_type"),
                    "record_id": value(row, "cellontology_id") or f"CellMarker2.0:{row_number}",
                    "reference": "https://bio-bigdata.hrbmu.edu.cn/CellMarker/",
                }
            )
    tmp.replace(output)


def install_msigdb_hallmark(output: Path, *, source_path: str | None, source_url: str | None, force: bool) -> dict[str, object]:
    source_path = source_path or os.environ.get("GENOMI_MSIGDB_HALLMARK_GMT")
    source_url = source_url or os.environ.get("GENOMI_MSIGDB_HALLMARK_GMT_URL")
    if output.is_file() and not force:
        return {"status": "cached", "output": str(output), "library": "msigdb-hallmark"}
    if source_url:
        return download_library_file(source_url, output, library="msigdb-hallmark", force=force)
    if not source_path:
        raise SystemExit(
            "msigdb-hallmark requires --msigdb-gmt, --msigdb-gmt-url, "
            "GENOMI_MSIGDB_HALLMARK_GMT, or GENOMI_MSIGDB_HALLMARK_GMT_URL."
        )
    source = Path(source_path).expanduser()
    if not source.is_file():
        raise SystemExit(f"MSigDB Hallmark GMT does not exist: {source}")
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, output)
    write_manifest(output, library="msigdb-hallmark", source_url=str(source), transform="Copied from official user-supplied MSigDB Hallmark GMT export.")
    return {"status": "completed", "output": str(output), "library": "msigdb-hallmark"}


def install_pharmcat(output: Path, *, version: str | None, force: bool) -> dict[str, object]:
    if output.is_file() and not force:
        return {"status": "cached", "output": str(output), "library": "pharmcat"}
    release = _fetch_pharmcat_release(version)
    asset = _select_pharmcat_jar_asset(release)
    if asset is None:
        raise SystemExit("PharmCAT release does not contain an all-in-one JAR asset.")
    summary = download_library_file(asset["browser_download_url"], output, library="pharmcat", force=force)
    write_manifest(
        output,
        library="pharmcat",
        source_url=str(asset["browser_download_url"]),
        transform=f"Downloaded {asset['name']} from PharmCAT release {release.get('tag_name') or 'unknown'}.",
    )
    summary["pharmcat_version"] = release.get("tag_name")
    summary["pharmcat_asset"] = asset["name"]
    return summary


def _fetch_pharmcat_release(version: str | None) -> dict[str, object]:
    url = (
        f"https://api.github.com/repos/PharmGKB/PharmCAT/releases/tags/{version}"
        if version
        else PHARMCAT_RELEASES_API_URL
    )
    request = urllib.request.Request(
        url,
        headers={"User-Agent": GENOMI_USER_AGENT, "Accept": "application/vnd.github+json"},
    )
    with urllib.request.urlopen(request, timeout=60) as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise SystemExit(f"PharmCAT releases API returned an unexpected payload from {url}")
    return payload


def _select_pharmcat_jar_asset(release: dict[str, object]) -> dict[str, object] | None:
    assets = release.get("assets") or []
    if not isinstance(assets, list):
        return None
    candidates = [asset for asset in assets if isinstance(asset, dict) and isinstance(asset.get("name"), str)]
    for asset in candidates:
        name = str(asset.get("name") or "")
        if name.startswith("pharmcat-") and name.endswith("-all.jar"):
            return asset
    for asset in candidates:
        if str(asset.get("name") or "").endswith(".jar"):
            return asset
    return None


def install_aligner_binary(
    *,
    install_dir: Path,
    binary_path: Path,
    library: str,
    archive_url: str,
    archive_sha256: str,
    binary_name: str,
    version: str,
    force: bool,
) -> dict[str, object]:
    """Download, verify, and extract a pinned aligner binary release.

    Currently fetches the linux_x64 tarball that minimap2 and bwa-mem2 ship
    on their GitHub releases. The tarballs contain a top-level directory
    (`minimap2-2.28_x64-linux/`, `bwa-mem2-2.2.1_x64-linux/`) with the
    binary at its root; we copy only the named binary into
    `<GENOMI_HOME>/tools/aligners/<aligner>/<binary>` so the runtime has a
    stable path independent of the upstream release filename.
    """

    manifest = install_dir / "manifest.json"
    if binary_path.is_file() and manifest.exists() and not force:
        return {
            "status": "cached",
            "output": str(binary_path),
            "library": library,
            "version": version,
        }
    if sys.platform != "linux":
        raise SystemExit(
            f"Aligner library {library} currently ships a linux x86_64 binary only; "
            f"detected sys.platform={sys.platform!r}. Install the aligner via your "
            f"system package manager (e.g. brew install {binary_name}) and place it on PATH."
        )

    install_dir.mkdir(parents=True, exist_ok=True)
    tarball_path = install_dir / f"{binary_name}-{version}.tar.bz2.partial"
    request = urllib.request.Request(archive_url, headers={"User-Agent": GENOMI_USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=300) as response, tarball_path.open("wb") as handle:
            shutil.copyfileobj(response, handle)
    except urllib.error.HTTPError as exc:
        tarball_path.unlink(missing_ok=True)
        raise SystemExit(
            f"Could not download {library} from {archive_url}: HTTP {exc.code}."
        ) from exc

    actual_sha = _sha256_file(tarball_path)
    if actual_sha.lower() != archive_sha256.lower():
        tarball_path.unlink(missing_ok=True)
        raise SystemExit(
            f"{library} tarball checksum mismatch: expected {archive_sha256}, got {actual_sha}. "
            f"If you are pointing the installer at a custom URL, also pass the matching --aligner-sha256."
        )

    extracted_binary: Path | None = None
    with tarfile.open(tarball_path, "r:bz2") as tar:
        for member in tar.getmembers():
            if not member.isfile():
                continue
            if Path(member.name).name != binary_name:
                continue
            member.name = binary_name  # flatten the leading release dir
            tar.extract(member, install_dir, filter="data")
            extracted_binary = install_dir / binary_name
            break
    tarball_path.unlink(missing_ok=True)
    if extracted_binary is None or not extracted_binary.is_file():
        raise SystemExit(
            f"{library} tarball at {archive_url} did not contain a {binary_name} binary."
        )
    extracted_binary.chmod(0o755)
    if extracted_binary != binary_path:
        binary_path.parent.mkdir(parents=True, exist_ok=True)
        extracted_binary.replace(binary_path)

    manifest.write_text(
        json.dumps(
            {
                "library": library,
                "binary": str(binary_path),
                "version": version,
                "source_url": archive_url,
                "source_sha256": archive_sha256,
                "downloaded_at": datetime.now(timezone.utc).isoformat(),
            },
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )
    return {
        "status": "completed",
        "output": str(binary_path),
        "library": library,
        "version": version,
    }


def write_manifest(output: Path, *, library: str, source_url: str, transform: str) -> None:
    manifest = {
        "library": library,
        "source_url": source_url,
        "output": str(output),
        "transform": transform,
        "downloaded_at": datetime.now(timezone.utc).isoformat(),
    }
    output.with_name(output.name + ".genomi-manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def run(command: list[str]) -> None:
    subprocess.run(command, check=True)
